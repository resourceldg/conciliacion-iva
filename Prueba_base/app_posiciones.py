"""
Posiciones IVA
==============
Subí el WPP (.xlsx) y el PDF DDJJ de ARCA → genera el Posiciones para ese período.

Flujo:
  1. Leer el período desde el PDF de ARCA
  2. Extraer los valores de ese período desde el WPP
  3. Inyectar los valores en el template Posiciones (ajustando la fecha de encabezado)
  4. Descargar el resultado
"""

import io
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

from conciliacion.posiciones import (
    CAMPOS_LABELS,
    MESES_ES,
    DDJJData,
    build_posiciones,
    extract_empresa_wpp,
    read_posiciones_pdf,
    read_wpp,
)

# ── Configuración ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Posiciones IVA",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

DATA_DIR = Path("data/posiciones")
TEMPLATE_PATH = DATA_DIR / "Posiciones.xlsx"

st.markdown("""
<style>
  thead th { background-color: #1e3a5f !important; color: white !important; }
</style>
""", unsafe_allow_html=True)

# ── Título ─────────────────────────────────────────────────────────────────────

st.title("📋 Posiciones IVA")
st.caption("Subí el WPP y el PDF de ARCA para generar el Posiciones del período indicado en el PDF.")

# ── Sidebar: carga de archivos ─────────────────────────────────────────────────

with st.sidebar:
    st.header("Archivos")

    wpp_upload = st.file_uploader(
        "WPP IVA (.xlsx)",
        type=["xlsx"],
        key="wpp_upload",
        help="Working Papers del IVA. Un sheet por mes (MM-YYYY).",
    )
    pdf_upload = st.file_uploader(
        "DDJJ ARCA (.pdf)",
        type=["pdf"],
        key="pdf_upload",
        help="F.2051 descargado de ARCA. Define el período a generar.",
    )

    wpp_source = wpp_upload or None
    pdf_source = pdf_upload or None

    st.divider()
    st.caption("Módulo independiente del Conciliador IVA.")

# ── Validación básica ──────────────────────────────────────────────────────────

if wpp_source is None:
    st.info("Cargá el archivo WPP (.xlsx) en el sidebar.")
    st.stop()

if pdf_source is None:
    st.info("Cargá el PDF de ARCA (.pdf) en el sidebar.")
    st.stop()

if not TEMPLATE_PATH.exists():
    st.error(f"Template no encontrado: `{TEMPLATE_PATH}`. Asegurate de que exista en `data/posiciones/`.")
    st.stop()

# ── Leer archivos ──────────────────────────────────────────────────────────────

def _read_bytes(src) -> bytes:
    if hasattr(src, "read"):
        b = src.read()
        if hasattr(src, "seek"):
            src.seek(0)
        return b
    return Path(src).read_bytes()


@st.cache_data(show_spinner="Leyendo WPP…")
def _load_wpp(b: bytes, name: str) -> DDJJData:
    return read_wpp(b)


@st.cache_data(show_spinner="Leyendo PDF ARCA…")
def _load_pdf(b: bytes, name: str) -> dict:
    return read_posiciones_pdf(b)


wpp_bytes = _read_bytes(wpp_source)
pdf_bytes = _read_bytes(pdf_source)

try:
    wpp_data: DDJJData = _load_wpp(wpp_bytes, getattr(wpp_source, "name", "wpp"))
    pdf_data: dict = _load_pdf(pdf_bytes, getattr(pdf_source, "name", "pdf"))
except Exception as e:
    st.error(f"Error leyendo archivos: {e}")
    st.stop()

if not wpp_data:
    st.error("No se encontraron hojas de meses (MM-YYYY) en el WPP.")
    st.stop()

if not pdf_data:
    st.error("No se pudo determinar el período desde el PDF de ARCA. Verificá que sea un F.2051 válido.")
    st.stop()

# ── Determinar período (lo define el PDF) ──────────────────────────────────────

(mes, anio) = max(pdf_data.keys(), key=lambda k: (k[1], k[0]))
periodo_label = f"{MESES_ES[mes]} {anio}"

if (mes, anio) not in wpp_data:
    st.error(
        f"El período **{periodo_label}** indicado por el PDF no está en el WPP. "
        f"Períodos disponibles: {', '.join(f'{MESES_ES[m]} {a}' for m, a in sorted(wpp_data))}"
    )
    st.stop()

# ── Resumen de lo que se va a generar ─────────────────────────────────────────

empresa = extract_empresa_wpp(wpp_bytes)

col1, col2, col3, col4 = st.columns(4)
col1.metric("Empresa", empresa or "—")
col2.metric("Período (del PDF)", periodo_label)
col3.metric("Períodos en WPP", len(wpp_data))
col4.metric("Template", TEMPLATE_PATH.name)

# ── Valores extraídos del WPP para ese período ────────────────────────────────

st.subheader(f"Valores WPP — {periodo_label}")
vals = wpp_data[(mes, anio)]
row_data = {label: vals.get(campo) for campo, label in CAMPOS_LABELS.items()}
df = pd.DataFrame([row_data])
fmt = {c: "{:,.2f}" for c in df.columns}
st.dataframe(df.style.format(fmt, na_rep="-"), use_container_width=True)

# ── Generar ────────────────────────────────────────────────────────────────────

st.divider()
st.subheader("Generar Posiciones")

col_btn, col_info = st.columns([1, 3])
with col_btn:
    generar = st.button("⚙️ Generar", type="primary", use_container_width=True)
with col_info:
    st.caption(
        f"Toma el template `{TEMPLATE_PATH.name}`, actualiza el encabezado a **{periodo_label}** "
        "y escribe los valores del WPP."
    )

if generar:
    with st.spinner("Generando Posiciones…"):
        # Cargar template y actualizar la fecha del encabezado al período del PDF
        tmpl_buf = io.BytesIO(TEMPLATE_PATH.read_bytes())
        wb_tmpl = openpyxl.load_workbook(tmpl_buf, data_only=False)
        ws_tmpl = wb_tmpl[next(s for s in wb_tmpl.sheetnames if "IVA" in s.upper())]

        # Actualizar empresa (fila 2, col B) desde el WPP
        if empresa:
            for row_cells in ws_tmpl.iter_rows(min_row=2, max_row=2):
                for cell in row_cells:
                    if isinstance(cell.value, str) and cell.value.strip():
                        cell.value = empresa
                        break

        # Actualizar la fecha en el encabezado de columna al período del PDF
        for row_cells in ws_tmpl.iter_rows():
            has_concepto = any(
                isinstance(c.value, str) and "CONCEPTO" in c.value.upper()
                for c in row_cells
            )
            if has_concepto:
                for c in row_cells:
                    if isinstance(c.value, datetime):
                        c.value = datetime(anio, mes, 1)
                break

        updated_tmpl = io.BytesIO()
        wb_tmpl.save(updated_tmpl)
        updated_tmpl.seek(0)

        # Escribir valores (solo el período del PDF)
        buf = io.BytesIO()
        try:
            warns = build_posiciones(updated_tmpl, {(mes, anio): vals}, buf)
            st.session_state["pos_bytes"] = buf.getvalue()
            st.session_state["pos_warns"] = warns
            st.session_state["pos_period"] = (mes, anio)
            st.session_state["pos_empresa"] = empresa
        except Exception as e:
            st.error(f"Error generando Posiciones: {e}")
            st.stop()

# ── Resultado ──────────────────────────────────────────────────────────────────

if st.session_state.get("pos_period") == (mes, anio) and "pos_bytes" in st.session_state:
    warns = st.session_state.get("pos_warns", {})
    real_warns = {
        p: m for p, m in warns.items()
        if not all("no encontrado en las columnas" in msg for msg in m)
    }

    if real_warns:
        with st.expander(f"⚠️ {len(real_warns)} advertencia(s)", expanded=True):
            for period, msgs in sorted(real_warns.items()):
                for msg in msgs:
                    st.warning(f"**{period}:** {msg}")
    else:
        st.success(f"Posiciones generado correctamente para {periodo_label}.")

    mes_str = MESES_ES[mes].lower()
    empresa_slug = st.session_state.get("pos_empresa", "").replace(" ", "_").upper()
    fname = f"Posiciones_IVA_{empresa_slug}_{mes_str}_{anio}.xlsx" if empresa_slug else f"Posiciones_IVA_{mes_str}_{anio}.xlsx"
    st.download_button(
        label=f"📥 Descargar {fname}",
        data=st.session_state["pos_bytes"],
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
