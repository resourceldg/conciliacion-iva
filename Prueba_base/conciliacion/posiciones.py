"""
Módulo Posiciones IVA
=====================
Lee el WPP (Working Papers IVA, una hoja por mes MM-YYYY) y actualiza
el archivo de Posiciones contables.

Flujo:
  1. read_wpp()       → extrae los valores DDJJ de cada hoja mensual
  2. build_posiciones() → escribe esos valores en el template Posiciones
"""

from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import IO

import openpyxl

# ── Configuración WPP ─────────────────────────────────────────────────────────

# Columna G (índice 6) = etiqueta, columna H (índice 7) = valor
_WPP_LABEL_IDX = 6
_WPP_VALUE_IDX = 7

_WPP_LABELS: dict[str, str] = {
    "DEBITO FISCAL":     "debito_fiscal",
    "CREDITO FISCAL":    "credito_fiscal",
    "S/F TEC. ANTERIOR": "saldo_tec_anterior",
    "ID ( A FAVOR )":   "impuesto_determinado",
    "RETENCIONES IVA":   "retenciones_iva",
    "PERCEPCIONES IVA":  "percepciones_iva",
    "S/F L.D. ANTERIOR": "saldo_ld_anterior",
    "A PAGAR (A FAVOR)": "saldo_final",
}

_SHEET_RE = re.compile(r"^(\d{2})-(\d{4})$")

# ── Tipos ─────────────────────────────────────────────────────────────────────

# { (mes, año): { campo: valor } }
DDJJData = dict[tuple[int, int], dict[str, float]]

# { "MM-YYYY": [advertencia, ...] }
UpdateWarnings = dict[str, list[str]]


# ── Leer WPP ─────────────────────────────────────────────────────────────────

def _parse_sheet_month(name: str) -> tuple[int, int] | None:
    m = _SHEET_RE.match(name.strip())
    return (int(m.group(1)), int(m.group(2))) if m else None


def _extract_sheet_values(ws) -> dict[str, float]:
    """Extrae los valores de la sección CONTROL DDJJ IVA de una hoja WPP."""
    found: dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) <= _WPP_VALUE_IDX:
            continue
        label = row[_WPP_LABEL_IDX]
        value = row[_WPP_VALUE_IDX]
        if not isinstance(label, str):
            continue
        key = label.strip()
        if key in _WPP_LABELS and isinstance(value, (int, float)):
            found[_WPP_LABELS[key]] = float(value)
    return found


def extract_empresa_wpp(source: Path | IO | bytes) -> str:
    """Extrae el nombre de la empresa del WPP (fila 4, col B, primera hoja mensual)."""
    if isinstance(source, bytes):
        source = BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    for sname in wb.sheetnames:
        if not _SHEET_RE.match(sname.strip()):
            continue
        ws = wb[sname]
        for row in ws.iter_rows(min_row=3, max_row=6, values_only=True):
            cell = row[1] if len(row) > 1 else None
            if isinstance(cell, str) and cell.strip():
                name = cell.strip()
                if name.upper().startswith("EMPRESA "):
                    name = name[8:].strip()
                return name
    return ""


def read_wpp(source: Path | IO | bytes) -> DDJJData:
    """
    Lee todas las hojas mensuales (MM-YYYY) del WPP.
    Acepta Path, file-like object o bytes.
    Retorna { (mes, año): { campo: valor } }.
    """
    if isinstance(source, bytes):
        source = BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True)
    result: DDJJData = {}
    for sname in wb.sheetnames:
        parsed = _parse_sheet_month(sname)
        if parsed is None:
            continue
        data = _extract_sheet_values(wb[sname])
        if data:
            result[parsed] = data
    return result


# ── Estructura del archivo Posiciones ─────────────────────────────────────────

class _Block:
    """Un bloque semestral (6 meses) dentro de una hoja de Posiciones."""

    def __init__(self, header_row: int):
        self.header_row = header_row
        # (mes, año) → columna 1-based en openpyxl
        self.month_cols: dict[tuple[int, int], int] = {}
        # label_upper → fila 1-based
        self.concept_rows: dict[str, int] = {}


def _scan_iva_sheet(ws) -> list[_Block]:
    """
    Escanea la hoja IVA del archivo Posiciones.
    Detecta bloques por la presencia de 'CONCEPTO' y fechas en la misma fila.
    """
    blocks: list[_Block] = []
    current: _Block | None = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        has_concepto = any(isinstance(c, str) and "CONCEPTO" in c.upper() for c in row)
        has_date = any(isinstance(c, datetime) for c in row)

        if has_concepto and has_date:
            blk = _Block(row_idx)
            for ci, cell in enumerate(row):
                if isinstance(cell, datetime):
                    # ci es 0-based; openpyxl usa 1-based
                    blk.month_cols[(cell.month, cell.year)] = ci + 1
            blocks.append(blk)
            current = blk
            continue

        if current is not None:
            # El concepto está en la columna D (índice 3, col 4 en openpyxl)
            concept_cell = row[3] if len(row) > 3 else None
            if isinstance(concept_cell, str) and concept_cell.strip():
                label = concept_cell.strip().upper()
                current.concept_rows[label] = row_idx

    return blocks


def _find_row(block: _Block, keyword: str, exclude: str = "") -> int | None:
    """
    Busca fila por substring (case-insensitive).
    Si exclude se indica, descarta filas que lo contengan.
    """
    kw = keyword.strip().upper()
    exc = exclude.strip().upper()
    # Intento exacto primero
    if kw in block.concept_rows:
        return block.concept_rows[kw]
    # Substring
    for label, row in block.concept_rows.items():
        if kw in label and (not exc or exc not in label):
            return row
    return None


# ── Actualizar Posiciones ─────────────────────────────────────────────────────

def build_posiciones(
    posiciones_source: Path | IO | bytes,
    wpp_data: DDJJData,
    output: Path | IO,
) -> UpdateWarnings:
    """
    Actualiza el archivo Posiciones con los valores extraídos del WPP.

    Para cada período en wpp_data escribe:
      D  IVA DEBITO FISCAL               ← debito_fiscal (abs)
      D  IVA LIBRE DISP. SDO. A FAVOR   ← abs(saldo_final) si saldo_final < 0
      H  IVA CREDITO FISCAL              ← credito_fiscal (abs)
      H  RETENCIONES IVA                 ← retenciones_iva (abs)
      H  PERCEPCIONES IVA                ← percepciones_iva (abs)
      H  IVA LIBRE DISP. SDO. A FAVOR (ant.) ← saldo_ld_anterior (abs)
      H  IVA A PAGAR                     ← saldo_final si saldo_final > 0

    Retorna advertencias por período (vacío = sin problemas).
    """
    if isinstance(posiciones_source, bytes):
        posiciones_source = BytesIO(posiciones_source)
    if isinstance(output, (str, Path)):
        output = Path(output)

    # data_only=True: lee valores en caché de fórmulas y los guarda como literales.
    # Evita que openpyxl invalide la caché de celdas con fórmula al guardar.
    wb = openpyxl.load_workbook(posiciones_source, data_only=True)

    iva_sheets = [s for s in wb.sheetnames if "IVA" in s.upper()]
    if not iva_sheets:
        raise ValueError("No se encontró hoja de IVA en el archivo Posiciones.")

    ws = wb[iva_sheets[0]]
    blocks = _scan_iva_sheet(ws)
    warnings: UpdateWarnings = {}

    for (mes, anio), vals in sorted(wpp_data.items()):
        key = f"{mes:02d}-{anio}"
        w: list[str] = []

        target = next((b for b in blocks if (mes, anio) in b.month_cols), None)
        if target is None:
            w.append(f"Período {key} no encontrado en las columnas de Posiciones.")
            warnings[key] = w
            continue

        col = target.month_cols[(mes, anio)]

        def set_cell(kw: str, value: float | None, exclude: str = "", force_abs: bool = True) -> None:
            if value is None:
                return
            row = _find_row(target, kw, exclude=exclude)
            if row is None:
                w.append(f"Concepto '{kw}' no encontrado en bloque {key}.")
                return
            ws.cell(row=row, column=col, value=abs(value) if force_abs else value)

        # Débito y crédito fiscal
        set_cell("IVA DEBITO FISCAL",  vals.get("debito_fiscal"))
        set_cell("IVA CREDITO FISCAL", vals.get("credito_fiscal"))

        # Retenciones y percepciones
        set_cell("RETENCIONES IVA",  vals.get("retenciones_iva"))
        set_cell("PERCEPCIONES IVA", vals.get("percepciones_iva"))

        # Saldo LD anterior → H IVA LIBRE DISP. SDO. A FAVOR (ant.)
        # Búsqueda doble para no confundir con 'IVA TECNICO SDO. A FAVOR (ant.)'
        sla = vals.get("saldo_ld_anterior")
        if sla is not None:
            row_ant = next(
                (r for lbl, r in target.concept_rows.items()
                 if "IVA LIBRE DISP" in lbl and "ANT" in lbl),
                None,
            )
            if row_ant:
                ws.cell(row=row_ant, column=col, value=abs(sla))
            else:
                w.append("'IVA LIBRE DISP. SDO. A FAVOR (ant.)' no encontrado.")

        # Saldo final
        sf = vals.get("saldo_final")
        if sf is not None:
            if sf < 0:
                # A favor → D IVA LIBRE DISP. SDO. A FAVOR (sin "ant.")
                row_favor = None
                for label, r in target.concept_rows.items():
                    if "IVA LIBRE DISP" in label and "ANT" not in label:
                        row_favor = r
                        break
                if row_favor:
                    ws.cell(row=row_favor, column=col, value=abs(sf))
                else:
                    w.append("'IVA LIBRE DISP. SDO. A FAVOR' (D) no encontrado.")
            else:
                set_cell("IVA A PAGAR", sf, force_abs=False)

        if w:
            warnings[key] = w

    if isinstance(output, Path):
        wb.save(output)
    else:
        wb.save(output)

    return warnings


# ── Helpers para UI ───────────────────────────────────────────────────────────

CAMPOS_LABELS: dict[str, str] = {
    "debito_fiscal":       "Débito Fiscal",
    "credito_fiscal":      "Crédito Fiscal (WPP)",
    "saldo_tec_anterior":  "Saldo Técnico Anterior",
    "impuesto_determinado": "Impuesto Determinado",
    "retenciones_iva":     "Retenciones IVA",
    "percepciones_iva":    "Percepciones IVA",
    "saldo_ld_anterior":   "S/F L.D. Anterior",
    "saldo_final":         "A Pagar / (A Favor)",
}

MESES_ES: dict[int, str] = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


# ── Leer valores actuales de Posiciones (para comparación) ───────────────────

# Mapeo de campo WPP → keyword de concepto en Posiciones (para comparar)
_CAMPO_TO_CONCEPTO: dict[str, tuple[str, str]] = {
    # campo_wpp: (keyword_a_buscar, exclude)
    "debito_fiscal":    ("IVA DEBITO FISCAL",  ""),
    "credito_fiscal":   ("IVA CREDITO FISCAL",  ""),
    "retenciones_iva":  ("RETENCIONES IVA",    ""),
    "percepciones_iva": ("PERCEPCIONES IVA",   ""),
    "saldo_ld_anterior": ("IVA LIBRE DISP",    "ANT"),   # excluye ant.
    "saldo_final":      ("IVA LIBRE DISP",     "ANT"),   # si a favor
}


def read_posiciones_iva(
    source: Path | IO | bytes,
) -> dict[tuple[int, int], dict[str, float | None]]:
    """
    Lee los valores actuales de la hoja IVA del archivo Posiciones.
    Retorna { (mes, año): { campo_wpp: valor_actual } } para comparar con WPP.
    """
    if isinstance(source, bytes):
        source = BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True)
    iva_sheets = [s for s in wb.sheetnames if "IVA" in s.upper()]
    if not iva_sheets:
        return {}

    ws = wb[iva_sheets[0]]
    blocks = _scan_iva_sheet(ws)
    result: dict[tuple[int, int], dict[str, float | None]] = {}

    for block in blocks:
        for (mes, anio), col in block.month_cols.items():
            vals: dict[str, float | None] = {}
            for campo, (kw, exc) in _CAMPO_TO_CONCEPTO.items():
                kw_up  = kw.strip().upper()
                exc_up = exc.strip().upper()
                row = next(
                    (r for lbl, r in block.concept_rows.items()
                     if kw_up in lbl and (not exc_up or exc_up not in lbl)),
                    None,
                )
                if row is not None:
                    v = ws.cell(row=row, column=col).value
                    vals[campo] = float(v) if isinstance(v, (int, float)) else None
                else:
                    vals[campo] = None
            # Saldo LD anterior (con ANT) — campo especial
            row_ant = next(
                (r for lbl, r in block.concept_rows.items()
                 if "IVA LIBRE DISP" in lbl and "ANT" in lbl),
                None,
            )
            v_ant = ws.cell(row=row_ant, column=col).value if row_ant else None
            vals["saldo_ld_anterior"] = float(v_ant) if isinstance(v_ant, (int, float)) else None

            result[(mes, anio)] = vals

    return result
